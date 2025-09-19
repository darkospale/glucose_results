#!/usr/bin/env python3
"""
Simplified Glucose Data Converter with Template Support and Date Filtering
Author: Your Assistant
Purpose: Converts CSV glucose data to formatted XLSX with single template
"""

import os
import sys
import csv
import json
import platform
import argparse
import configparser
import shutil
from datetime import datetime, timedelta
from pathlib import Path
from typing import List, Dict, Tuple, Optional, Any
import openpyxl
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet


class ExportTracker:
    """Tracks last export date for incremental exports"""
    
    def __init__(self, tracker_file: str = None):
        """Initialize export tracker"""
        self.tracker_file = tracker_file or str(Path.home() / '.glucose_last_export.json')
        self.last_exports = self._load_tracker()
    
    def _load_tracker(self) -> Dict:
        """Load export dates from file"""
        if os.path.exists(self.tracker_file):
            try:
                with open(self.tracker_file, 'r') as f:
                    return json.load(f)
            except:
                return {}
        return {}
    
    def save_tracker(self):
        """Save export dates to file"""
        with open(self.tracker_file, 'w') as f:
            json.dump(self.last_exports, f, indent=2, default=str)
    
    def get_last_export_date(self, source_file: str) -> Optional[datetime]:
        """Get the last export date for a specific source file"""
        if source_file in self.last_exports:
            try:
                return datetime.fromisoformat(self.last_exports[source_file])
            except:
                return None
        return None
    
    def update_export(self, source_file: str, latest_date: datetime):
        """Update the export date for a file"""
        self.last_exports[source_file] = latest_date.isoformat()
        self.save_tracker()


class SimplifiedGlucoseConverter:
    """Simplified converter with template support and date filtering"""
    
    # Color definitions for glucose level ranges
    COLORS = {
        'low': 'E6F2FF',      # Pale blue for < 4.0
        'high': 'FFCCCC',     # Red for > 11.9
        'very_high': 'E6D9FF'  # Light purple for > 17.9
    }
    
    def __init__(self, config_file: Optional[str] = None):
        """Initialize converter with optional config file"""
        self.config = self._load_config(config_file)
        self.export_tracker = ExportTracker()
        self.template_path = self._get_template_path()
    
    def _get_template_path(self) -> Optional[Path]:
        """Get path to user's template if it exists"""
        # Check for template in config directory
        template_dir = Path.home() / '.glucose_converter'
        template_dir.mkdir(exist_ok=True)
        
        template_file = template_dir / 'template.xlsx'
        if template_file.exists():
            return template_file
        
        # Check for template in current directory
        local_template = Path('glucose_template.xlsx')
        if local_template.exists():
            return local_template
        
        return None
    
    def save_template(self, template_file: str) -> bool:
        """Save a template file for future use"""
        template_dir = Path.home() / '.glucose_converter'
        template_dir.mkdir(exist_ok=True)
        
        dest_path = template_dir / 'template.xlsx'
        try:
            shutil.copy2(template_file, dest_path)
            self.template_path = dest_path
            print(f"✅ Template saved to: {dest_path}")
            return True
        except Exception as e:
            print(f"❌ Failed to save template: {e}")
            return False
    
    def _load_config(self, config_file: Optional[str]) -> Dict:
        """Load configuration from file or use defaults"""
        config = {
            'output_folder': None,  # None means same as input
            'auto_open': False,
            'date_format': '%d.%m.%Y %H:%M',
            'low_threshold': 4.0,
            'high_threshold': 11.9,
            'very_high_threshold': 17.9,
            'incremental_export': True,
            'date_filter_enabled': False,
            'date_filter_days': 30
        }
        
        if config_file and os.path.exists(config_file):
            parser = configparser.ConfigParser()
            parser.read(config_file)
            
            if 'Settings' in parser:
                settings = parser['Settings']
                for key in config:
                    if key in settings:
                        if key in ['auto_open', 'incremental_export', 'date_filter_enabled']:
                            config[key] = settings.getboolean(key)
                        elif key in ['low_threshold', 'high_threshold', 'very_high_threshold']:
                            config[key] = float(settings[key])
                        elif key == 'date_filter_days':
                            config[key] = int(settings[key])
                        else:
                            config[key] = settings[key] if settings[key] else None
        
        return config
    
    def read_csv(self, csv_path: str, start_date: Optional[datetime] = None, 
                 end_date: Optional[datetime] = None) -> List[Dict]:
        """Read CSV file with optional date filtering"""
        data = []
        
        with open(csv_path, 'r', encoding='utf-8-sig') as csvfile:
            reader = csv.DictReader(csvfile)
            
            for row in reader:
                date_str = row.get('Date and Time', '').strip()
                glucose_str = row.get('Readings [mmol/L]', '').strip()
                
                if date_str and glucose_str:
                    try:
                        # Parse date: "DD.M.YY. H:MM" format
                        parts = date_str.split(' ')
                        date_part = parts[0].rstrip('.')
                        time_part = parts[1] if len(parts) > 1 else '00:00'
                        
                        # Split date components
                        date_components = date_part.split('.')
                        if len(date_components) == 3:
                            day = int(date_components[0])
                            month = int(date_components[1])
                            year = int(date_components[2])
                            # Convert 2-digit year to 4-digit
                            if year < 100:
                                year = 2000 + year
                            
                            # Parse time
                            time_components = time_part.split(':')
                            hour = int(time_components[0])
                            minute = int(time_components[1]) if len(time_components) > 1 else 0
                            
                            # Create datetime object
                            dt = datetime(year, month, day, hour, minute)
                            
                            # Apply date filtering
                            if start_date and dt < start_date:
                                continue
                            if end_date and dt > end_date:
                                continue
                            
                            data.append({
                                'datetime': dt,
                                'glucose': float(glucose_str),
                                'meal_marker': row.get('Meal Marker', ''),
                                'notes': row.get('Notes', ''),
                                'activity': row.get('Activity', ''),
                                'meal': row.get('Meal[g]', ''),
                                'medication': row.get('Medication', ''),
                                'location': row.get('Location', '')
                            })
                    except (ValueError, IndexError) as e:
                        print(f"Warning: Could not parse row with date '{date_str}': {e}")
                        continue
        
        # Sort by datetime
        data.sort(key=lambda x: x['datetime'])
        return data
    
    def get_cell_color(self, glucose_value: float) -> Optional[str]:
        """Determine cell background color based on glucose value"""
        if glucose_value < self.config['low_threshold']:
            return self.COLORS['low']
        elif glucose_value > self.config['very_high_threshold']:
            return self.COLORS['very_high']
        elif glucose_value > self.config['high_threshold']:
            return self.COLORS['high']
        return None
    
    def apply_template_formatting(self, ws: Worksheet, template_ws: Worksheet, data_rows: int):
        """Apply formatting from template to worksheet"""
        # Copy column widths
        for col in template_ws.column_dimensions:
            ws.column_dimensions[col].width = template_ws.column_dimensions[col].width
        
        # Copy row heights for headers
        if 1 in template_ws.row_dimensions:
            ws.row_dimensions[1].height = template_ws.row_dimensions[1].height
        
        # Copy header styles
        for col in range(1, min(9, template_ws.max_column + 1)):
            template_cell = template_ws.cell(row=1, column=col)
            ws_cell = ws.cell(row=1, column=col)
            
            if template_cell.font:
                ws_cell.font = Font(
                    name=template_cell.font.name,
                    size=template_cell.font.size,
                    bold=template_cell.font.bold,
                    italic=template_cell.font.italic,
                    color=template_cell.font.color
                )
            
            if template_cell.fill and template_cell.fill.patternType:
                ws_cell.fill = PatternFill(
                    start_color=template_cell.fill.start_color.rgb if template_cell.fill.start_color else 'FFFFFF',
                    end_color=template_cell.fill.end_color.rgb if template_cell.fill.end_color else 'FFFFFF',
                    fill_type=template_cell.fill.patternType
                )
            
            if template_cell.alignment:
                ws_cell.alignment = Alignment(
                    horizontal=template_cell.alignment.horizontal,
                    vertical=template_cell.alignment.vertical,
                    wrap_text=template_cell.alignment.wrap_text
                )
    
    def create_xlsx(self, data: List[Dict], output_path: str):
        """Create XLSX using template if available or default formatting"""
        
        # Try to load template
        template_wb = None
        if self.template_path and self.template_path.exists():
            try:
                template_wb = load_workbook(str(self.template_path))
                print(f"📋 Using template from: {self.template_path}")
            except Exception as e:
                print(f"⚠️ Could not load template: {e}")
        
        if template_wb:
            wb = template_wb
            ws = wb.active
            # Clear existing data rows (keep header)
            for row in range(2, ws.max_row + 1):
                for col in range(1, ws.max_column + 1):
                    ws.cell(row=row, column=col).value = None
        else:
            wb = Workbook()
            ws = wb.active
            ws.title = "Glucose Readings"
        
        # Define headers
        headers = [
            'Date and Time',
            'Glucose [mmol/L]',
            'Meal Marker',
            'Notes',
            'Activity',
            'Meal [g]',
            'Medication',
            'Location'
        ]
        
        # Apply default formatting if no template
        if not template_wb:
            header_font = Font(bold=True, size=11)
            header_fill = PatternFill(start_color='D3D3D3', end_color='D3D3D3', fill_type='solid')
            header_alignment = Alignment(horizontal='center', vertical='center')
            
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            # Write headers
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                cell.border = thin_border
        else:
            # Update headers if using template
            for col, header in enumerate(headers, 1):
                ws.cell(row=1, column=col, value=header)
        
        # Write data rows
        for row_idx, row_data in enumerate(data, 2):
            # Date and Time column
            date_cell = ws.cell(
                row=row_idx, 
                column=1, 
                value=row_data['datetime'].strftime(self.config['date_format'])
            )
            
            # Glucose value column with color coding
            glucose_cell = ws.cell(row=row_idx, column=2, value=row_data['glucose'])
            
            # Apply color based on glucose level
            color = self.get_cell_color(row_data['glucose'])
            if color:
                glucose_cell.fill = PatternFill(
                    start_color=color, 
                    end_color=color, 
                    fill_type='solid'
                )
            
            # Other columns
            other_values = [
                row_data['meal_marker'],
                row_data['notes'],
                row_data['activity'],
                row_data['meal'],
                row_data['medication'],
                row_data['location']
            ]
            
            for col_idx, value in enumerate(other_values, 3):
                ws.cell(row=row_idx, column=col_idx, value=value if value else '')
            
            # Apply borders if not using template
            if not template_wb:
                for col in range(1, 9):
                    ws.cell(row=row_idx, column=col).border = thin_border
        
        # Auto-adjust column widths if no template
        if not template_wb:
            for column in ws.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)
                
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                
                adjusted_width = min(max(max_length + 2, 10), 50)
                ws.column_dimensions[column_letter].width = adjusted_width
        
        # Add statistics
        self._add_statistics(ws, data, len(data) + 3)
        
        # Save the workbook
        wb.save(output_path)
        print(f"✅ Successfully created XLSX file: {output_path}")
        
        # Auto-open if configured
        if self.config['auto_open']:
            self._open_file(output_path)
    
    def _add_statistics(self, ws: Worksheet, data: List[Dict], start_row: int):
        """Add statistics summary to worksheet"""
        if not data:
            return
        
        glucose_values = [d['glucose'] for d in data]
        avg_glucose = sum(glucose_values) / len(glucose_values)
        min_glucose = min(glucose_values)
        max_glucose = max(glucose_values)
        
        # Count readings in different ranges
        low_count = sum(1 for v in glucose_values if v < self.config['low_threshold'])
        normal_count = sum(1 for v in glucose_values if self.config['low_threshold'] <= v <= self.config['high_threshold'])
        high_count = sum(1 for v in glucose_values if self.config['high_threshold'] < v <= self.config['very_high_threshold'])
        very_high_count = sum(1 for v in glucose_values if v > self.config['very_high_threshold'])
        
        # Date range info
        date_range = f"{data[0]['datetime'].strftime('%d.%m.%Y')} - {data[-1]['datetime'].strftime('%d.%m.%Y')}"
        
        ws.cell(row=start_row, column=1, value='STATISTICS').font = Font(bold=True, size=12)
        
        stats = [
            ('Date Range:', date_range),
            ('Total Readings:', len(data)),
            ('Average Glucose:', f'{avg_glucose:.1f} mmol/L'),
            ('Minimum Glucose:', f'{min_glucose:.1f} mmol/L'),
            ('Maximum Glucose:', f'{max_glucose:.1f} mmol/L'),
            ('', ''),
            ('RANGE DISTRIBUTION:', ''),
            (f'Low (< {self.config["low_threshold"]} mmol/L):', f'{low_count} ({low_count/len(data)*100:.1f}%)' if data else '0 (0%)'),
            (f'Normal ({self.config["low_threshold"]}-{self.config["high_threshold"]} mmol/L):', f'{normal_count} ({normal_count/len(data)*100:.1f}%)' if data else '0 (0%)'),
            (f'High ({self.config["high_threshold"]}-{self.config["very_high_threshold"]} mmol/L):', f'{high_count} ({high_count/len(data)*100:.1f}%)' if data else '0 (0%)'),
            (f'Very High (> {self.config["very_high_threshold"]} mmol/L):', f'{very_high_count} ({very_high_count/len(data)*100:.1f}%)' if data else '0 (0%)')
        ]
        
        for idx, (label, value) in enumerate(stats, 1):
            label_cell = ws.cell(row=start_row + idx, column=1, value=label)
            value_cell = ws.cell(row=start_row + idx, column=2, value=value)
            
            if 'STATISTICS' in label or 'DISTRIBUTION' in label:
                label_cell.font = Font(bold=True)
    
    def _open_file(self, filepath: str):
        """Open file with cross-platform support"""
        import subprocess
        
        system = platform.system()
        
        try:
            if system == 'Windows':
                os.startfile(filepath)
            elif system == 'Darwin':  # macOS
                subprocess.run(['open', filepath], check=True)
            else:  # Linux and others
                subprocess.run(['xdg-open', filepath], check=True)
        except Exception as e:
            print(f"Could not auto-open file: {e}")
    
    def convert(self, csv_path: str, output_path: Optional[str] = None,
                start_date: Optional[datetime] = None,
                end_date: Optional[datetime] = None,
                incremental: bool = None) -> str:
        """Main conversion with date filtering"""
        
        # Validate input file
        if not os.path.exists(csv_path):
            raise FileNotFoundError(f"CSV file not found: {csv_path}")
        
        # Handle incremental export
        if incremental is None:
            incremental = self.config.get('incremental_export', False)
        
        if incremental:
            last_export = self.export_tracker.get_last_export_date(csv_path)
            if last_export and not start_date:
                start_date = last_export
                print(f"📅 Using incremental export from: {start_date.strftime('%d.%m.%Y %H:%M')}")
        
        # Apply date filter if configured
        if self.config.get('date_filter_enabled') and not end_date:
            days = self.config.get('date_filter_days', 30)
            end_date = datetime.now()
            if not start_date:
                start_date = end_date - timedelta(days=days)
        
        # Determine output path
        if not output_path:
            csv_path_obj = Path(csv_path)
            base_name = csv_path_obj.stem
            
            if self.config['output_folder']:
                output_dir = Path(self.config['output_folder'])
                output_dir.mkdir(parents=True, exist_ok=True)
            else:
                output_dir = csv_path_obj.parent
            
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            output_path = output_dir / f"{base_name}_formatted_{timestamp}.xlsx"
        
        # Read CSV data with filtering
        print(f"📖 Reading CSV file: {csv_path}")
        data = self.read_csv(csv_path, start_date, end_date)
        
        if not data:
            print("⚠️ No data found in the specified date range")
            return None
        
        print(f"✅ Found {len(data)} glucose readings")
        if start_date or end_date:
            date_range = []
            if start_date:
                date_range.append(f"from {start_date.strftime('%d.%m.%Y')}")
            if end_date:
                date_range.append(f"to {end_date.strftime('%d.%m.%Y')}")
            print(f"📅 Date range: {' '.join(date_range)}")
        
        # Create XLSX file
        print(f"📝 Creating formatted XLSX file...")
        self.create_xlsx(data, str(output_path))
        
        # Update export tracker if incremental
        if incremental and data:
            latest_date = max(d['datetime'] for d in data)
            self.export_tracker.update_export(csv_path, latest_date)
            print(f"📅 Saved last export date: {latest_date.strftime('%d.%m.%Y %H:%M')}")
        
        return str(output_path)


def find_latest_csv(folder_path: str) -> Optional[str]:
    """Find the most recent Contour CSV file in a folder"""
    folder = Path(folder_path)
    csv_files = list(folder.glob('ContourCSVReport*.csv'))
    
    if not csv_files:
        return None
    
    csv_files.sort(key=lambda x: x.stat().st_mtime, reverse=True)
    return str(csv_files[0])


def get_downloads_folder() -> Path:
    """Get the downloads folder path for the current platform"""
    system = platform.system()
    
    if system == 'Windows':
        try:
            import winreg
            sub_key = r'SOFTWARE\Microsoft\Windows\CurrentVersion\Explorer\Shell Folders'
            with winreg.OpenKey(winreg.HKEY_CURRENT_USER, sub_key) as key:
                downloads = winreg.QueryValueEx(key, '{374DE290-123F-4565-9164-39C4925E467B}')[0]
            return Path(downloads)
        except:
            pass
    
    # Default for all platforms
    downloads = Path.home() / 'Downloads'
    if not downloads.exists():
        downloads = Path.home() / 'downloads'
    return downloads


def main():
    """Simplified command-line interface"""
    parser = argparse.ArgumentParser(
        description='Simplified Glucose Converter with Date Filtering',
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  %(prog)s --auto-detect                    # Auto-detect latest CSV
  %(prog)s input.csv --incremental          # Only export new data since last export
  %(prog)s input.csv --last-days 7          # Export last 7 days only
  %(prog)s --upload-template template.xlsx  # Upload your XLSX template
        """
    )
    
    parser.add_argument('input', nargs='?', help='Path to CSV file')
    parser.add_argument('-o', '--output', help='Output XLSX file path')
    parser.add_argument('-c', '--config', help='Path to configuration file')
    parser.add_argument('--auto-detect', action='store_true', help='Auto-detect latest CSV')
    parser.add_argument('--create-config', action='store_true', help='Create sample config')
    
    # Template option (simplified)
    parser.add_argument('--upload-template', metavar='FILE',
                       help='Upload XLSX template for formatting')
    
    # Date filtering options
    parser.add_argument('--incremental', action='store_true', 
                       help='Only export data since last export')
    parser.add_argument('--start-date', help='Start date (DD.MM.YYYY)')
    parser.add_argument('--end-date', help='End date (DD.MM.YYYY)')
    parser.add_argument('--last-days', type=int, help='Export last N days only')
    
    # Reset incremental tracking
    parser.add_argument('--reset-tracker', action='store_true',
                       help='Reset incremental export tracking')
    
    args = parser.parse_args()
    
    # Create converter
    converter = SimplifiedGlucoseConverter(config_file=args.config)
    
    # Handle template upload
    if args.upload_template:
        if not os.path.exists(args.upload_template):
            print(f"Error: Template file not found: {args.upload_template}")
            return 1
        
        if converter.save_template(args.upload_template):
            print("Template will be used for all future conversions")
        return 0
    
    # Create sample config
    if args.create_config:
        config_content = """[Settings]
# Output folder path (leave empty to use same folder as input)
output_folder = 

# Automatically open the file after conversion
auto_open = false

# Glucose thresholds (in mmol/L)
low_threshold = 4.0
high_threshold = 11.9
very_high_threshold = 17.9

# Export settings
incremental_export = true
date_filter_enabled = false
date_filter_days = 30
"""
        config_path = 'glucose_config.ini'
        with open(config_path, 'w') as f:
            f.write(config_content)
        print(f"✅ Created configuration file: {config_path}")
        return 0
    
    # Determine input file
    input_file = args.input
    
    if args.auto_detect or not input_file:
        downloads = get_downloads_folder()
        print(f"🔍 Looking for latest CSV in {downloads}...")
        input_file = find_latest_csv(str(downloads))
        
        if not input_file:
            print("❌ No Contour CSV files found in Downloads folder")
            return 1
        
        print(f"✅ Found: {input_file}")
    
    # Reset tracker if requested
    if args.reset_tracker:
        converter.export_tracker.last_exports = {}
        converter.export_tracker.save_tracker()
        print("✅ Reset incremental export tracking")
        return 0
    
    # Parse date filters
    start_date = None
    end_date = None
    
    if args.start_date:
        try:
            start_date = datetime.strptime(args.start_date, '%d.%m.%Y')
        except ValueError:
            print(f"Error: Invalid start date format. Use DD.MM.YYYY")
            return 1
    
    if args.end_date:
        try:
            end_date = datetime.strptime(args.end_date, '%d.%m.%Y')
        except ValueError:
            print(f"Error: Invalid end date format. Use DD.MM.YYYY")
            return 1
    
    if args.last_days:
        end_date = datetime.now()
        start_date = end_date - timedelta(days=args.last_days)
    
    # Run conversion
    try:
        output_file = converter.convert(
            input_file, 
            args.output,
            start_date=start_date,
            end_date=end_date,
            incremental=args.incremental
        )
        
        if output_file:
            print(f"\n🎉 Conversion complete!")
            print(f"📄 Output saved to: {output_file}")
        
    except Exception as e:
        print(f"\n❌ Error: {e}")
        import traceback
        traceback.print_exc()
        return 1
    
    return 0


if __name__ == '__main__':
    sys.exit(main())