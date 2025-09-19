#!/usr/bin/env python3
"""
Glucose Data Converter for Contour Plus CSV exports to formatted XLSX
Author: Your Assistant
Purpose: Converts CSV glucose data to color-coded XLSX format
"""

import os
import sys
import csv
import argparse
import configparser
from datetime import datetime
from pathlib import Path
from typing import List, Dict, Tuple, Optional
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter


class GlucoseConverter:
    """Main converter class for glucose CSV to XLSX conversion"""
    
    # Color definitions for glucose level ranges
    COLORS = {
        'low': 'E6F2FF',      # Pale blue for < 4.0
        'high': 'FFCCCC',     # Red for > 11.9
        'very_high': 'E6D9FF'  # Light purple for > 17.9
    }
    
    def __init__(self, config_file: Optional[str] = None):
        """Initialize converter with optional config file"""
        self.config = self._load_config(config_file)
        
    def _load_config(self, config_file: Optional[str]) -> Dict:
        """Load configuration from file or use defaults"""
        config = {
            'output_folder': None,  # None means same as input
            'auto_open': False,
            'date_format': '%d.%m.%Y %H:%M',
            'low_threshold': 4.0,
            'high_threshold': 11.9,
            'very_high_threshold': 17.9
        }
        
        if config_file and os.path.exists(config_file):
            parser = configparser.ConfigParser()
            parser.read(config_file)
            
            if 'Settings' in parser:
                settings = parser['Settings']
                if 'output_folder' in settings:
                    config['output_folder'] = settings['output_folder']
                if 'auto_open' in settings:
                    config['auto_open'] = settings.getboolean('auto_open')
                if 'low_threshold' in settings:
                    config['low_threshold'] = float(settings['low_threshold'])
                if 'high_threshold' in settings:
                    config['high_threshold'] = float(settings['high_threshold'])
                if 'very_high_threshold' in settings:
                    config['very_high_threshold'] = float(settings['very_high_threshold'])
        
        return config
    
    def read_csv(self, csv_path: str) -> List[Dict]:
        """Read CSV file and return data as list of dictionaries"""
        data = []
        
        with open(csv_path, 'r', encoding='utf-8-sig') as csvfile:
            # Skip BOM if present
            reader = csv.DictReader(csvfile)
            
            for row in reader:
                # Parse the date format from CSV (format: "14.5.25. 6:31")
                date_str = row.get('Date and Time', '').strip()
                glucose_str = row.get('Readings [mmol/L]', '').strip()
                
                if date_str and glucose_str:
                    try:
                        # Parse date: "DD.M.YY. H:MM" format
                        # Need to handle the year format (25 -> 2025)
                        parts = date_str.split(' ')
                        date_part = parts[0].rstrip('.')
                        time_part = parts[1] if len(parts) > 1 else '00:00'
                        
                        # Split date components
                        date_components = date_part.split('.')
                        if len(date_components) == 3:
                            day = int(date_components[0])
                            month = int(date_components[1])
                            year = int(date_components[2])
                            # Convert 2-digit year to 4-digit
                            if year < 100:
                                year = 2000 + year
                            
                            # Parse time
                            time_components = time_part.split(':')
                            hour = int(time_components[0])
                            minute = int(time_components[1]) if len(time_components) > 1 else 0
                            
                            # Create datetime object
                            dt = datetime(year, month, day, hour, minute)
                            
                            data.append({
                                'datetime': dt,
                                'glucose': float(glucose_str),
                                'meal_marker': row.get('Meal Marker', ''),
                                'notes': row.get('Notes', ''),
                                'activity': row.get('Activity', ''),
                                'meal': row.get('Meal[g]', ''),
                                'medication': row.get('Medication', ''),
                                'location': row.get('Location', '')
                            })
                    except (ValueError, IndexError) as e:
                        print(f"Warning: Could not parse row with date '{date_str}': {e}")
                        continue
        
        return data
    
    def get_cell_color(self, glucose_value: float) -> Optional[str]:
        """Determine cell background color based on glucose value"""
        if glucose_value < self.config['low_threshold']:
            return self.COLORS['low']
        elif glucose_value > self.config['very_high_threshold']:
            return self.COLORS['very_high']
        elif glucose_value > self.config['high_threshold']:
            return self.COLORS['high']
        return None
    
    def create_xlsx(self, data: List[Dict], output_path: str):
        """Create formatted XLSX file from glucose data"""
        
        # Create workbook and worksheet
        wb = Workbook()
        ws = wb.active
        ws.title = "Glucose Readings"
        
        # Define headers
        headers = [
            'Date and Time',
            'Glucose [mmol/L]',
            'Meal Marker',
            'Notes',
            'Activity',
            'Meal [g]',
            'Medication',
            'Location'
        ]
        
        # Style definitions
        header_font = Font(bold=True, size=11)
        header_fill = PatternFill(start_color='D3D3D3', end_color='D3D3D3', fill_type='solid')
        header_alignment = Alignment(horizontal='center', vertical='center')
        
        # Border style
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Write headers
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border
        
        # Write data rows
        for row_idx, row_data in enumerate(data, 2):
            # Date and Time column
            date_cell = ws.cell(
                row=row_idx, 
                column=1, 
                value=row_data['datetime'].strftime(self.config['date_format'])
            )
            date_cell.border = thin_border
            date_cell.alignment = Alignment(horizontal='left', vertical='center')
            
            # Glucose value column with color coding
            glucose_cell = ws.cell(row=row_idx, column=2, value=row_data['glucose'])
            glucose_cell.border = thin_border
            glucose_cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # Apply color based on glucose level
            color = self.get_cell_color(row_data['glucose'])
            if color:
                glucose_cell.fill = PatternFill(
                    start_color=color, 
                    end_color=color, 
                    fill_type='solid'
                )
            
            # Other columns
            other_values = [
                row_data['meal_marker'],
                row_data['notes'],
                row_data['activity'],
                row_data['meal'],
                row_data['medication'],
                row_data['location']
            ]
            
            for col_idx, value in enumerate(other_values, 3):
                cell = ws.cell(row=row_idx, column=col_idx, value=value if value else '')
                cell.border = thin_border
                cell.alignment = Alignment(horizontal='left', vertical='center')
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            
            # Set column width with minimum and maximum limits
            adjusted_width = min(max(max_length + 2, 10), 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Add summary statistics at the bottom
        last_row = len(data) + 3
        
        # Calculate statistics
        glucose_values = [d['glucose'] for d in data]
        avg_glucose = sum(glucose_values) / len(glucose_values) if glucose_values else 0
        min_glucose = min(glucose_values) if glucose_values else 0
        max_glucose = max(glucose_values) if glucose_values else 0
        
        # Count readings in different ranges
        low_count = sum(1 for v in glucose_values if v < self.config['low_threshold'])
        normal_count = sum(1 for v in glucose_values if self.config['low_threshold'] <= v <= self.config['high_threshold'])
        high_count = sum(1 for v in glucose_values if self.config['high_threshold'] < v <= self.config['very_high_threshold'])
        very_high_count = sum(1 for v in glucose_values if v > self.config['very_high_threshold'])
        
        # Write statistics
        stats_start_row = last_row
        
        ws.cell(row=stats_start_row, column=1, value='STATISTICS').font = Font(bold=True, size=12)
        
        stats = [
            ('Total Readings:', len(data)),
            ('Average Glucose:', f'{avg_glucose:.1f} mmol/L'),
            ('Minimum Glucose:', f'{min_glucose:.1f} mmol/L'),
            ('Maximum Glucose:', f'{max_glucose:.1f} mmol/L'),
            ('', ''),  # Empty row
            ('RANGE DISTRIBUTION:', ''),
            (f'Low (< {self.config["low_threshold"]} mmol/L):', f'{low_count} ({low_count/len(data)*100:.1f}%)'),
            (f'Normal ({self.config["low_threshold"]}-{self.config["high_threshold"]} mmol/L):', f'{normal_count} ({normal_count/len(data)*100:.1f}%)'),
            (f'High ({self.config["high_threshold"]}-{self.config["very_high_threshold"]} mmol/L):', f'{high_count} ({high_count/len(data)*100:.1f}%)'),
            (f'Very High (> {self.config["very_high_threshold"]} mmol/L):', f'{very_high_count} ({very_high_count/len(data)*100:.1f}%)')
        ]
        
        for idx, (label, value) in enumerate(stats, 1):
            label_cell = ws.cell(row=stats_start_row + idx, column=1, value=label)
            value_cell = ws.cell(row=stats_start_row + idx, column=2, value=value)
            
            if 'STATISTICS' in label or 'DISTRIBUTION' in label:
                label_cell.font = Font(bold=True)
        
        # Save the workbook
        wb.save(output_path)
        print(f"✅ Successfully created XLSX file: {output_path}")
        
        # Auto-open if configured
        if self.config['auto_open']:
            self._open_file(output_path)
    
    def _open_file(self, filepath: str):
        """Open the file with the default system application"""
        import platform
        import subprocess
        
        if platform.system() == 'Windows':
            os.startfile(filepath)
        elif platform.system() == 'Darwin':  # macOS
            subprocess.run(['open', filepath])
        else:  # Linux
            subprocess.run(['xdg-open', filepath])
    
    def convert(self, csv_path: str, output_path: Optional[str] = None) -> str:
        """Main conversion method"""
        
        # Validate input file
        if not os.path.exists(csv_path):
            raise FileNotFoundError(f"CSV file not found: {csv_path}")
        
        # Determine output path
        if not output_path:
            # Generate output filename based on input
            csv_path_obj = Path(csv_path)
            base_name = csv_path_obj.stem
            
            # Use configured output folder or same as input
            if self.config['output_folder']:
                output_dir = Path(self.config['output_folder'])
                output_dir.mkdir(parents=True, exist_ok=True)
            else:
                output_dir = csv_path_obj.parent
            
            # Create output filename with timestamp
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            output_path = output_dir / f"{base_name}_formatted_{timestamp}.xlsx"
        
        # Read CSV data
        print(f"📖 Reading CSV file: {csv_path}")
        data = self.read_csv(csv_path)
        
        if not data:
            raise ValueError("No valid data found in CSV file")
        
        print(f"✅ Found {len(data)} glucose readings")
        
        # Create XLSX file
        print(f"📝 Creating formatted XLSX file...")
        self.create_xlsx(data, str(output_path))
        
        return str(output_path)


def find_latest_csv(folder_path: str) -> Optional[str]:
    """Find the most recent Contour CSV file in a folder"""
    folder = Path(folder_path)
    csv_files = list(folder.glob('ContourCSVReport*.csv'))
    
    if not csv_files:
        return None
    
    # Sort by modification time
    csv_files.sort(key=lambda x: x.stat().st_mtime, reverse=True)
    return str(csv_files[0])


def main():
    """Command-line interface for the converter"""
    parser = argparse.ArgumentParser(
        description='Convert Contour Plus glucose CSV to formatted XLSX'
    )
    
    parser.add_argument(
        'input',
        nargs='?',
        help='Path to CSV file (or leave empty to auto-detect in Downloads)'
    )
    
    parser.add_argument(
        '-o', '--output',
        help='Output XLSX file path'
    )
    
    parser.add_argument(
        '-c', '--config',
        help='Path to configuration file'
    )
    
    parser.add_argument(
        '--auto-detect',
        action='store_true',
        help='Automatically find latest CSV in Downloads folder'
    )
    
    parser.add_argument(
        '--create-config',
        action='store_true',
        help='Create a sample configuration file'
    )
    
    args = parser.parse_args()
    
    # Create sample config if requested
    if args.create_config:
        config_content = """[Settings]
# Output folder path (leave empty to use same folder as input)
output_folder = 

# Automatically open the file after conversion
auto_open = false

# Glucose thresholds (in mmol/L)
low_threshold = 4.0
high_threshold = 11.9
very_high_threshold = 17.9
"""
        config_path = 'glucose_config.ini'
        with open(config_path, 'w') as f:
            f.write(config_content)
        print(f"✅ Created sample configuration file: {config_path}")
        print("Edit this file to customize your settings.")
        return
    
    # Determine input file
    input_file = args.input
    
    if args.auto_detect or not input_file:
        # Try to find in Downloads folder
        downloads = Path.home() / 'Downloads'
        print(f"🔍 Looking for latest CSV in {downloads}...")
        input_file = find_latest_csv(str(downloads))
        
        if not input_file:
            print("❌ No Contour CSV files found in Downloads folder")
            print("Please specify the CSV file path")
            return 1
        
        print(f"✅ Found: {input_file}")
    
    # Create converter and run conversion
    try:
        converter = GlucoseConverter(config_file=args.config)
        output_file = converter.convert(input_file, args.output)
        print(f"\n🎉 Conversion complete!")
        print(f"📄 Output saved to: {output_file}")
        
    except Exception as e:
        print(f"\n❌ Error: {e}")
        return 1
    
    return 0


if __name__ == '__main__':
    sys.exit(main())