#!/usr/bin/env python3
"""
Test Suite for Glucose Converter
Tests all features including template support, date filtering, and cross-platform compatibility
"""

import unittest
import tempfile
import os
import shutil
import csv
import json
from datetime import datetime, timedelta
from pathlib import Path
import platform
from unittest.mock import patch, MagicMock

# Import modules to test
from glucose_converter_enhanced import (
    EnhancedGlucoseConverter,
    ExportTracker,
    TemplateManager,
    find_latest_csv,
    get_downloads_folder
)

import openpyxl


class TestExportTracker(unittest.TestCase):
    """Test the ExportTracker class"""
    
    def setUp(self):
        """Create temporary tracker file"""
        self.temp_dir = tempfile.mkdtemp()
        self.tracker_file = os.path.join(self.temp_dir, 'tracker.json')
        self.tracker = ExportTracker(self.tracker_file)
    
    def tearDown(self):
        """Clean up temp files"""
        shutil.rmtree(self.temp_dir)
    
    def test_empty_tracker(self):
        """Test empty tracker initialization"""
        self.assertEqual(self.tracker.history, {})
        self.assertIsNone(self.tracker.get_last_export_date('nonexistent.csv'))
    
    def test_update_export(self):
        """Test updating export record"""
        test_file = 'test.csv'
        test_date = datetime(2025, 1, 15, 10, 30)
        
        self.tracker.update_export(test_file, test_date)
        
        # Check history updated
        self.assertIn(test_file, self.tracker.history)
        self.assertEqual(
            self.tracker.history[test_file]['last_export'],
            test_date.isoformat()
        )
        
        # Check file saved
        self.assertTrue(os.path.exists(self.tracker_file))
        
        # Check retrieval
        retrieved_date = self.tracker.get_last_export_date(test_file)
        self.assertEqual(retrieved_date, test_date)
    
    def test_load_existing_history(self):
        """Test loading existing history file"""
        # Create history file
        history_data = {
            'test.csv': {
                'last_export': datetime(2025, 1, 10).isoformat(),
                'updated_at': datetime.now().isoformat()
            }
        }
        
        with open(self.tracker_file, 'w') as f:
            json.dump(history_data, f)
        
        # Load tracker
        new_tracker = ExportTracker(self.tracker_file)
        
        # Check loaded correctly
        self.assertEqual(len(new_tracker.history), 1)
        self.assertIn('test.csv', new_tracker.history)


class TestTemplateManager(unittest.TestCase):
    """Test the TemplateManager class"""
    
    def setUp(self):
        """Create temporary template directory"""
        self.temp_dir = tempfile.mkdtemp()
        self.manager = TemplateManager(self.temp_dir)
    
    def tearDown(self):
        """Clean up temp files"""
        shutil.rmtree(self.temp_dir)
    
    def test_empty_templates(self):
        """Test empty template list"""
        templates = self.manager.list_templates()
        self.assertEqual(templates, [])
    
    def test_save_and_list_template(self):
        """Test saving and listing templates"""
        # Create a dummy XLSX file
        wb = openpyxl.Workbook()
        temp_file = os.path.join(self.temp_dir, 'temp.xlsx')
        wb.save(temp_file)
        
        # Save as template
        self.manager.save_template(temp_file, 'test_template')
        
        # Check template saved
        templates = self.manager.list_templates()
        self.assertIn('test_template', templates)
        
        # Check template path
        template_path = self.manager.get_template_path('test_template')
        self.assertIsNotNone(template_path)
        self.assertTrue(template_path.exists())
    
    def test_load_template(self):
        """Test loading a template"""
        # Create and save template
        wb = openpyxl.Workbook()
        ws = wb.active
        ws['A1'] = 'Test Header'
        
        temp_file = os.path.join(self.temp_dir, 'temp.xlsx')
        wb.save(temp_file)
        
        self.manager.save_template(temp_file, 'test_template')
        
        # Load template
        loaded_wb = self.manager.load_template('test_template')
        self.assertIsNotNone(loaded_wb)
        self.assertEqual(loaded_wb.active['A1'].value, 'Test Header')
    
    def test_nonexistent_template(self):
        """Test loading nonexistent template"""
        template_path = self.manager.get_template_path('nonexistent')
        self.assertIsNone(template_path)
        
        loaded = self.manager.load_template('nonexistent')
        self.assertIsNone(loaded)


class TestGlucoseConverter(unittest.TestCase):
    """Test the main converter functionality"""
    
    def setUp(self):
        """Setup test environment"""
        self.temp_dir = tempfile.mkdtemp()
        self.converter = EnhancedGlucoseConverter()
        
        # Create sample CSV data
        self.csv_file = os.path.join(self.temp_dir, 'test_glucose.csv')
        self.create_sample_csv()
    
    def tearDown(self):
        """Clean up"""
        shutil.rmtree(self.temp_dir)
    
    def create_sample_csv(self, num_days=10):
        """Create sample CSV file with glucose data"""
        headers = [
            '#', 'Date and Time', 'Readings [mmol/L]', 'Meal Marker',
            'Data Source', 'Notes', 'Activity', 'Meal[g]', 'Medication', 'Location'
        ]
        
        data = []
        base_date = datetime.now() - timedelta(days=num_days)
        
        for i in range(num_days * 4):  # 4 readings per day
            date = base_date + timedelta(hours=i*6)
            date_str = f"{date.day}.{date.month}.{date.year % 100}. {date.hour}:{date.minute:02d}"
            
            # Vary glucose values
            glucose = 5.5 + (i % 10) * 1.5  # Range from 5.5 to 20.5
            
            data.append([
                str(i+1),
                date_str,
                str(glucose),
                'Before Meal' if i % 2 == 0 else 'After Meal',
                'Meter',
                '',
                '',
                '',
                '',
                ''
            ])
        
        with open(self.csv_file, 'w', newline='', encoding='utf-8-sig') as f:
            writer = csv.writer(f)
            writer.writerow(headers)
            writer.writerows(data)
    
    def test_read_csv(self):
        """Test CSV reading functionality"""
        data = self.converter.read_csv(self.csv_file)
        
        self.assertIsInstance(data, list)
        self.assertGreater(len(data), 0)
        
        # Check data structure
        first_row = data[0]
        self.assertIn('datetime', first_row)
        self.assertIn('glucose', first_row)
        self.assertIn('meal_marker', first_row)
        
        # Check data types
        self.assertIsInstance(first_row['datetime'], datetime)
        self.assertIsInstance(first_row['glucose'], float)
    
    def test_date_filtering(self):
        """Test date range filtering"""
        # Read all data
        all_data = self.converter.read_csv(self.csv_file)
        
        # Read with date filter
        end_date = datetime.now()
        start_date = end_date - timedelta(days=3)
        
        filtered_data = self.converter.read_csv(
            self.csv_file,
            start_date=start_date,
            end_date=end_date
        )
        
        # Check filtered data is subset
        self.assertLess(len(filtered_data), len(all_data))
        
        # Check all dates within range
        for row in filtered_data:
            self.assertGreaterEqual(row['datetime'], start_date)
            self.assertLessEqual(row['datetime'], end_date)
    
    def test_glucose_color_coding(self):
        """Test glucose level color determination"""
        # Test low glucose
        color = self.converter.get_cell_color(3.5)
        self.assertEqual(color, self.converter.COLORS['low'])
        
        # Test normal glucose
        color = self.converter.get_cell_color(7.0)
        self.assertIsNone(color)
        
        # Test high glucose
        color = self.converter.get_cell_color(15.0)
        self.assertEqual(color, self.converter.COLORS['high'])
        
        # Test very high glucose
        color = self.converter.get_cell_color(20.0)
        self.assertEqual(color, self.converter.COLORS['very_high'])
    
    def test_xlsx_creation(self):
        """Test XLSX file creation"""
        data = self.converter.read_csv(self.csv_file)
        output_file = os.path.join(self.temp_dir, 'output.xlsx')
        
        self.converter.create_xlsx_with_template(data, output_file)
        
        # Check file created
        self.assertTrue(os.path.exists(output_file))
        
        # Load and verify XLSX
        wb = openpyxl.load_workbook(output_file)
        ws = wb.active
        
        # Check headers
        self.assertEqual(ws['A1'].value, 'Date and Time')
        self.assertEqual(ws['B1'].value, 'Glucose [mmol/L]')
        
        # Check data rows
        self.assertGreater(ws.max_row, len(data))  # Data + headers + stats
        
        # Check glucose value formatting
        for row in range(2, len(data) + 2):
            glucose_cell = ws.cell(row=row, column=2)
            self.assertIsNotNone(glucose_cell.value)
            
            # Check color coding applied
            if glucose_cell.value < 4.0:
                self.assertIsNotNone(glucose_cell.fill.start_color)
    
    def test_template_application(self):
        """Test applying template to output"""
        # Create a template
        template_wb = openpyxl.Workbook()
        template_ws = template_wb.active
        
        # Set custom formatting in template
        template_ws['A1'].font = openpyxl.styles.Font(bold=True, size=14)
        template_ws.column_dimensions['A'].width = 20
        
        # Save template
        template_file = os.path.join(self.temp_dir, 'template.xlsx')
        template_wb.save(template_file)
        
        # Save to template manager
        self.converter.template_manager.save_template(template_file, 'test_template')
        
        # Convert with template
        data = self.converter.read_csv(self.csv_file)
        output_file = os.path.join(self.temp_dir, 'output_with_template.xlsx')
        
        self.converter.create_xlsx_with_template(data, output_file, 'test_template')
        
        # Verify template applied
        wb = openpyxl.load_workbook(output_file)
        ws = wb.active
        
        # Check formatting preserved
        self.assertTrue(ws['A1'].font.bold)
        self.assertEqual(ws['A1'].font.size, 14)
    
    def test_incremental_export(self):
        """Test incremental export functionality"""
        # First export
        output1 = os.path.join(self.temp_dir, 'output1.xlsx')
        result1 = self.converter.convert(self.csv_file, output1)
        
        self.assertIsNotNone(result1)
        
        # Get last export date
        last_export = self.converter.export_tracker.get_last_export_date(self.csv_file)
        self.assertIsNotNone(last_export)
        
        # Create new CSV with additional data
        self.create_sample_csv(num_days=15)  # More data
        
        # Second incremental export
        output2 = os.path.join(self.temp_dir, 'output2.xlsx')
        result2 = self.converter.convert(self.csv_file, output2, incremental=True)
        
        # Load both files and compare
        wb1 = openpyxl.load_workbook(output1)
        wb2 = openpyxl.load_workbook(output2)
        
        # Second file should have fewer rows (only new data)
        self.assertLess(wb2.active.max_row, wb1.active.max_row)
    
    def test_cross_platform_paths(self):
        """Test cross-platform path handling"""
        # Test path creation works on current platform
        test_path = Path.home() / 'test_folder'
        
        # Verify Path operations work
        self.assertIsInstance(test_path, Path)
        self.assertTrue(str(test_path))
        
        # Test downloads folder detection
        downloads = get_downloads_folder()
        self.assertIsInstance(downloads, Path)
    
    def test_config_loading(self):
        """Test configuration file loading"""
        config_file = os.path.join(self.temp_dir, 'test_config.ini')
        
        config_content = """[Settings]
output_folder = /test/output
auto_open = true
low_threshold = 3.5
high_threshold = 12.0
very_high_threshold = 18.0
incremental_export = true
date_filter_days = 7
"""
        
        with open(config_file, 'w') as f:
            f.write(config_content)
        
        # Create converter with config
        converter = EnhancedGlucoseConverter(config_file)
        
        # Check config loaded
        self.assertEqual(converter.config['output_folder'], '/test/output')
        self.assertTrue(converter.config['auto_open'])
        self.assertEqual(converter.config['low_threshold'], 3.5)
        self.assertEqual(converter.config['high_threshold'], 12.0)
        self.assertEqual(converter.config['very_high_threshold'], 18.0)
        self.assertTrue(converter.config['incremental_export'])
        self.assertEqual(converter.config['date_filter_days'], 7)


class TestUtilityFunctions(unittest.TestCase):
    """Test utility functions"""
    
    def test_find_latest_csv(self):
        """Test finding latest CSV file"""
        temp_dir = tempfile.mkdtemp()
        
        try:
            # Create multiple CSV files with different timestamps
            files = []
            for i in range(3):
                file_path = os.path.join(temp_dir, f'ContourCSVReport_{i}.csv')
                with open(file_path, 'w') as f:
                    f.write('test')
                files.append(file_path)
                
                # Modify file time
                mod_time = datetime.now().timestamp() - (i * 3600)
                os.utime(file_path, (mod_time, mod_time))
            
            # Find latest
            latest = find_latest_csv(temp_dir)
            
            self.assertIsNotNone(latest)
            self.assertTrue(latest.endswith('ContourCSVReport_0.csv'))
            
        finally:
            shutil.rmtree(temp_dir)
    
    def test_find_no_csv(self):
        """Test when no CSV files found"""
        temp_dir = tempfile.mkdtemp()
        
        try:
            latest = find_latest_csv(temp_dir)
            self.assertIsNone(latest)
        finally:
            shutil.rmtree(temp_dir)
    
    @patch('platform.system')
    def test_downloads_folder_windows(self, mock_system):
        """Test downloads folder detection on Windows"""
        mock_system.return_value = 'Windows'
        
        # Mock winreg for Windows
        with patch('glucose_converter_enhanced.winreg') as mock_winreg:
            mock_key = MagicMock()
            mock_winreg.OpenKey.return_value.__enter__.return_value = mock_key
            mock_winreg.QueryValueEx.return_value = ('C:\\Users\\Test\\Downloads', 1)
            
            downloads = get_downloads_folder()
            
            self.assertEqual(str(downloads), 'C:\\Users\\Test\\Downloads')
    
    @patch('platform.system')
    def test_downloads_folder_linux_mac(self, mock_system):
        """Test downloads folder detection on Linux/Mac"""
        mock_system.return_value = 'Linux'
        
        downloads = get_downloads_folder()
        
        # Should be home/Downloads or home/downloads
        self.assertTrue(str(downloads).endswith('Downloads') or 
                       str(downloads).endswith('downloads'))


class TestIntegration(unittest.TestCase):
    """Integration tests for complete workflow"""
    
    def setUp(self):
        """Setup test environment"""
        self.temp_dir = tempfile.mkdtemp()
    
    def tearDown(self):
        """Clean up"""
        shutil.rmtree(self.temp_dir)
    
    def test_complete_workflow(self):
        """Test complete conversion workflow"""
        # Create CSV
        csv_file = os.path.join(self.temp_dir, 'ContourCSVReport_test.csv')
        headers = [
            '#', 'Date and Time', 'Readings [mmol/L]', 'Meal Marker',
            'Data Source', 'Notes', 'Activity', 'Meal[g]', 'Medication', 'Location'
        ]
        
        data = [
            ['1', '1.1.25. 8:00', '3.5', 'Fasting', 'Meter', '', '', '', '', ''],  # Low
            ['2', '1.1.25. 12:00', '8.0', 'Before Meal', 'Meter', '', '', '', '', ''],  # Normal
            ['3', '1.1.25. 18:00', '15.0', 'After Meal', 'Meter', '', '', '', '', ''],  # High
            ['4', '1.1.25. 22:00', '22.0', 'No mark', 'Meter', '', '', '', '', ''],  # Very high
        ]
        
        with open(csv_file, 'w', newline='', encoding='utf-8-sig') as f:
            writer = csv.writer(f)
            writer.writerow(headers)
            writer.writerows(data)
        
        # Create converter
        converter = EnhancedGlucoseConverter()
        
        # Convert
        output_file = os.path.join(self.temp_dir, 'output.xlsx')
        result = converter.convert(csv_file, output_file)
        
        # Verify output
        self.assertIsNotNone(result)
        self.assertTrue(os.path.exists(output_file))
        
        # Load and check XLSX
        wb = openpyxl.load_workbook(output_file)
        ws = wb.active
        
        # Check all data present
        self.assertEqual(ws['B2'].value, 3.5)  # First glucose value
        self.assertEqual(ws['B3'].value, 8.0)  # Second glucose value
        self.assertEqual(ws['B4'].value, 15.0)  # Third glucose value
        self.assertEqual(ws['B5'].value, 22.0)  # Fourth glucose value
        
        # Check colors applied
        low_cell = ws['B2']
        high_cell = ws['B4']
        very_high_cell = ws['B5']
        
        # Check fills are applied (color-coded)
        self.assertIsNotNone(low_cell.fill.start_color)
        self.assertIsNotNone(high_cell.fill.start_color)
        self.assertIsNotNone(very_high_cell.fill.start_color)


def run_tests():
    """Run all tests"""
    # Create test suite
    loader = unittest.TestLoader()
    suite = unittest.TestSuite()
    
    # Add all test classes
    suite.addTests(loader.loadTestsFromTestCase(TestExportTracker))
    suite.addTests(loader.loadTestsFromTestCase(TestTemplateManager))
    suite.addTests(loader.loadTestsFromTestCase(TestGlucoseConverter))
    suite.addTests(loader.loadTestsFromTestCase(TestUtilityFunctions))
    suite.addTests(loader.loadTestsFromTestCase(TestIntegration))
    
    # Run tests
    runner = unittest.TextTestRunner(verbosity=2)
    result = runner.run(suite)
    
    # Print summary
    print("\n" + "="*60)
    print("TEST SUMMARY")
    print("="*60)
    print(f"Tests run: {result.testsRun}")
    print(f"Failures: {len(result.failures)}")
    print(f"Errors: {len(result.errors)}")
    print(f"Success rate: {((result.testsRun - len(result.failures) - len(result.errors)) / result.testsRun * 100):.1f}%")
    
    return result.wasSuccessful()


if __name__ == '__main__':
    success = run_tests()
    exit(0 if success else 1)